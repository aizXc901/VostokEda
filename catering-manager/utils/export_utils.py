"""
Утилиты для экспорта данных
"""

import csv
import json
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import Config
from models import *


class ExportUtils:
    """Утилиты для экспорта данных в различные форматы"""

    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: Path) -> bool:
        """Экспорт данных в CSV"""
        try:
            if not data:
                return False

            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = data[0].keys()
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)

            return True
        except Exception as e:
            print(f"Ошибка экспорта в CSV: {e}")
            return False

    @staticmethod
    def export_to_json(data: List[Dict[str, Any]], filename: Path) -> bool:
        """Экспорт данных в JSON"""
        try:
            # Конвертируем Decimal в строки для JSON
            def decimal_default(obj):
                if isinstance(obj, Decimal):
                    return str(obj)
                elif isinstance(obj, (datetime, date)):
                    return obj.isoformat()
                raise TypeError

            with open(filename, 'w', encoding='utf-8') as jsonfile:
                json.dump(data, jsonfile, ensure_ascii=False, indent=2, default=decimal_default)

            return True
        except Exception as e:
            print(f"Ошибка экспорта в JSON: {e}")
            return False

    @staticmethod
    def export_expense_report_to_excel(event_summary: EventSummary, filename: Path) -> bool:
        """Экспорт отчета по расходам в Excel"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчет по расходам"

            # Стили
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            money_format = '#,##0.00'
            percent_format = '0.0%'

            # Заголовок
            ws.merge_cells('A1:E1')
            ws['A1'] = f"Отчет по расходам на мероприятие: {event_summary.event.name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")

            # Информация о мероприятии
            ws['A3'] = "Мероприятие:"
            ws['B3'] = event_summary.event.name

            ws['A4'] = "Дата проведения:"
            ws['B4'] = Formatters.format_date(event_summary.event.event_date)

            ws['A5'] = "Общий бюджет:"
            ws['B5'] = Formatters.format_currency(event_summary.event.budget)

            ws['A6'] = "Общие расходы:"
            ws['B6'] = Formatters.format_currency(event_summary.total_amount)

            ws['A7'] = "Использовано бюджета:"
            ws['B7'] = Formatters.format_percentage(event_summary.budget_utilization)

            # Таблица расходов по категориям
            ws['A9'] = "Категория затрат"
            ws['B9'] = "План, руб"
            ws['C9'] = "Факт, руб"
            ws['D9'] = "Отклонение, руб"
            ws['E9'] = "% выполнения"

            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws[f'{col}9']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            row = 10
            for item in event_summary.categories_summary:
                ws[f'A{row}'] = item.category_name
                ws[f'B{row}'] = float(item.planned_amount)
                ws[f'B{row}'].number_format = money_format

                ws[f'C{row}'] = float(item.actual_amount)
                ws[f'C{row}'].number_format = money_format

                deviation = item.actual_amount - item.planned_amount
                ws[f'D{row}'] = float(deviation)
                ws[f'D{row}'].number_format = money_format

                percentage = item.percentage / 100
                ws[f'E{row}'] = percentage
                ws[f'E{row}'].number_format = percent_format

                # Цветовое выделение превышения бюджета
                if percentage > 1.0:  # Превышение 100%
                    ws[f'E{row}'].font = Font(color="FF0000", bold=True)

                row += 1

            # Итоговая строка
            ws[f'A{row}'] = "ИТОГО"
            ws[f'A{row}'].font = Font(bold=True)

            ws[f'B{row}'] = f"=SUM(B10:B{row - 1})"
            ws[f'B{row}'].number_format = money_format
            ws[f'B{row}'].font = Font(bold=True)

            ws[f'C{row}'] = f"=SUM(C10:C{row - 1})"
            ws[f'C{row}'].number_format = money_format
            ws[f'C{row}'].font = Font(bold=True)

            # Настройка ширины колонок
            column_widths = {'A': 30, 'B': 15, 'C': 15, 'D': 15, 'E': 15}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Сохранение
            wb.save(filename)
            return True

        except Exception as e:
            print(f"Ошибка экспорта в Excel: {e}")
            return False

    @staticmethod
    def prepare_event_data_for_export(event: Event, orders: List[Order]) -> Dict[str, Any]:
        """Подготовка данных мероприятия для экспорта"""
        order_details = []
        for order in orders:
            for item in order.items:
                order_details.append({
                    'Номер заказа': order.order_number,
                    'Дата заказа': Formatters.format_date(order.order_date),
                    'Позиция': item.nomenclature.name if item.nomenclature else '',
                    'Поставщик': item.supplier.name if item.supplier else '',
                    'Количество': float(item.quantity),
                    'Цена': Formatters.format_currency(item.unit_price),
                    'Стоимость': Formatters.format_currency(item.total_price),
                    'Примечания': item.notes
                })

        return {
            'Мероприятие': event.name,
            'Дата': Formatters.format_date(event.event_date),
            'Время': Formatters.format_time(event.start_time),
            'Количество гостей': event.guests_count,
            'Бюджет': Formatters.format_currency(event.budget),
            'Статус': event.status,
            'Место': event.location,
            'Ответственный': event.responsible_person,
            'Количество заказов': len(orders),
            'Общая сумма заказов': Formatters.format_currency(
                sum(order.total_amount for order in orders)
            ),
            'Детали заказов': order_details
        }
